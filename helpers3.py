import os
from openpyxl import load_workbook
from config import UPLOAD_FOLDER

for path, directory, files in os.walk(UPLOAD_FOLDER):
    #print (path, directory, files)
    for file in files:
        if file[-4:] == "xlsx" and "DRAS_" in file:
            print(file)
            wb = load_workbook(path + file)
            ws = wb.active
            for row in ws.iter_rows(min_row = 2):
                print(row[13].value)

    