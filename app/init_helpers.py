import openpyxl
from app import db
from app.models import (Mocmodel, Unitmodel, Dedocmodel,
                Splitofworks, Disciplinedras, Drascommentsheet,
                Drasissuetype, Drasactionrequired, Drascomment,
                Drasrevision, Drasdocument, Tagdiscipline)
from app.comments.helpers import get_data_from_cs, get_fake_data_from_cs, get_fake_data_from_cs2
from config import UPLOAD_FOLDER
import uuid
from flask_appbuilder.filemanager import FileManager
from flask import current_app
import random

path = UPLOAD_FOLDER + 'init/OperatingCenter.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active

path = UPLOAD_FOLDER + 'init/DisciplineTag.xlsx'
wbtag = openpyxl.load_workbook(path)
wstag = wbtag.active

def add_tags():
    session = db.session
    session.query(Tagdiscipline).delete()

    for row in wstag.iter_rows(min_row=2):
        print(row[0].value,row[1].value)
        tag = Tagdiscipline(name=row[0].value,
                            start=row[1].value,
                            finish=row[2].value,
                            created_by_fk='1',
                            changed_by_fk='1')
        session.add(tag)
    session.commit()

#add_tags()


def add_actionRequired():
    session = db.session
    session.query(Drasactionrequired).delete()

    ac_list = ['Action Required Type 1','Action Required Type 2','Action Required Type 3','Action Required Type 4','Action Required Type 5']
    
    for ac in ac_list:
        new_ac = Drasactionrequired(created_by_fk='1',
                        changed_by_fk='1',
                        name=random.choice(ac_list))
        session.add(new_ac)
    session.commit()

def add_issueType():
    session = db.session
    session.query(Drasissuetype).delete()

    issue_list = ['Issue Type 1','Issue Type 2','Issue Type 3','Issue Type 4','Issue Type 5']
    
    for issue in issue_list:
        print('random choice', random.choice(issue_list))
        new_issue = Drasissuetype(created_by_fk='1',
                        changed_by_fk='1',
                        name=random.choice(issue_list))
        session.add(new_issue)
    session.commit()

def add_discipline():
    session = db.session
    session.query(Disciplinedras).delete()

    disc_list = ['Process','Piping','Civil','Electrical','Instrument']

    for dsc in disc_list:
        new_dsc = Disciplinedras(created_by_fk='1',
                        changed_by_fk='1',
                        name=dsc.upper())
        session.add(new_dsc)
    session.commit()    

def add_moc():
    session = db.session
    session.query(Mocmodel).delete()

    main_op_set = set()
    for row in worksheet.iter_rows(min_row=2):
        main_op_set.add(row[1].value.strip())

    print(main_op_set)
    for moc in main_op_set:
        m = Mocmodel(name=moc, created_by_fk='1',changed_by_fk='1')
        print(m)
        session.add(m)
    session.commit()
    #result = session.query(Mocmodel)
    #return result   

def add_oc():
    session = db.session
    session.query(Dedocmodel).delete()

    oc_set = set()

    for row in worksheet.iter_rows(min_row=2):
        moc = row[1].value.strip()
        dedoc = row[2].value.strip()

        
        process = row[8].value.strip()
        piping = row[9].value.strip()
        civil = row[10].value.strip()
        electrical = row[11].value.strip()
        instrument = row[12].value.strip()
        
        list_d = [dedoc,process,piping,civil,electrical, instrument]
        print(moc, dedoc) 
        for e in list_d:
            #oc_set.add(e)
        
            xmoc = session.query(Mocmodel).filter(Mocmodel.name == moc).first()
            print(e) 
            if xmoc:
                xdoc = session.query(Dedocmodel).filter(Dedocmodel.moc_id == xmoc.id, Dedocmodel.name == e).first() 
                
                if xdoc is None:
                    new_dedoc = Dedocmodel(moc=xmoc, name=e, created_by_fk='1',changed_by_fk='1')
                    session.add(new_dedoc)
  
        session.commit()

def add_unit():
    session = db.session
    session.query(Unitmodel).delete()

    for row in worksheet.iter_rows(min_row=2):
        code = row[0].value.split(" - ")[0] 
        name = row[0].value.split(" - ")[1]
        xmoc = row[1].value.strip()
        xdedoc = row[2].value.strip() 

        moc = session.query(Mocmodel).filter(Mocmodel.name == xmoc).first()
        dedoc = session.query(Dedocmodel).filter(Dedocmodel.name == xdedoc).first()

        
        unit = Unitmodel(code=code, 
                    name=name,
                    moc=moc,
                    dedoc=dedoc, 
                    created_by_fk='1', 
                    changed_by_fk='1')
        
        session.add(unit)
        #print(unit,name)
    
    session.commit()


        # 
     
def splitOfWorks():
    session = db.session
    session.query(Splitofworks).delete()

    for row in worksheet.iter_rows(min_row=2):
        moc = row[1].value
        dedoc = row[2].value
        # Process,Piping,Civil,Electrical,Instrument
        process = (row[8].value,'Process')
        piping = (row[9].value, 'Piping')
        civil = (row[10].value, 'Civil')
        electrical = (row[11].value, 'Electrical')
        instrument = (row[12].value, 'Instrument')
        
        funit = row[0].value.split(" - ")[0]
        unit = session.query(Unitmodel).filter(Unitmodel.code == funit).first()
        
        list_d = [process,piping,civil,electrical, instrument]
        
        for e in list_d:
            discipline = session.query(Disciplinedras).filter(Disciplinedras.name == e[1].upper()).first()
            dedoc = session.query(Dedocmodel).filter(Dedocmodel.name == e[0]).first()
            sow = Splitofworks(unitmodel=unit,
                                disciplinedras=discipline,
                                oc = dedoc,
                                created_by_fk='1',
                                changed_by_fk='1')
            session.add(sow)
    session.commit()

def init_dras():
    session = db.session

    session.query(Drascomment).delete()
    session.query(Drascommentsheet).delete()
    session.query(Drasrevision).delete()
    session.query(Drasdocument).delete()

    session.query(Drasactionrequired).delete()
    session.query(Drasissuetype).delete()

    session.query(Splitofworks).delete()
    session.query(Disciplinedras).delete()
    session.query(Unitmodel).delete()
    session.query(Dedocmodel).delete()
    session.query(Mocmodel).delete()
    
    print('Initialization Start from:', path)
    print('')
    
    print('               -----  Action Required Init...')
    add_actionRequired()
    
    print('               -----  Issu Type Init...')
    add_issueType()


    print('               -----  MOC Init...')
    add_moc()


    print('               -----  DED OC Init...')
    add_oc()
    
    print('               -----  Discipline Init...')
    add_discipline()

    print('               -----  Unit Init...')
    add_unit()

    print('               -----  Split Of Works Init...')
    splitOfWorks()
    
    print('                ok    Init is Done')
  

#init_dras()   

##
##  Test
## 
#  
'''
from random import random, randint
session = db.session
fakeUnitList = session.query(Unitmodel).all()

def fakeDoc(revision,stage):
    
    
    #unit = str(randint(1,60)).zfill(2)
    unit = fakeUnitList[randint(1,len(fakeUnitList)-1)].code
    doctype = ['DW','PID','MOM']
    serial = str(randint(1,3)).zfill(4)
    sheet = str(randint(1,3)).zfill(2)
    #revision = ['A','A0','B','C']
    #stage = ['Y','Y1','Y2','YF']
    revision = revision
    stage = stage
    doc = "-".join(('DRAS_2544',
                unit,
                doctype[randint(0,len(doctype)-1)],
                serial,
                sheet + "_" + revision + stage+ ".xlsx"))
         
    return doc

#fakeDoc(90000)
# 
#  
import random
from datetime import datetime, timedelta

#done

# or a function
def gen_datetime(min_year=2017, max_year=datetime.now().year):
    # generate a datetime in format yyyy-mm-dd hh:mm:ss.000000
    start = datetime(min_year, 1, 1, 00, 00, 00)
    years = max_year - min_year + 1
    end = start + timedelta(days=365 * years)
    return start + (end - start) * random.random()

#print(randomDate("1/1/2017 1:30 PM", "1/1/2019 4:50 AM", random.random()))

import glob, os
from shutil import copyfile
from datetime import datetime
from app import app

def fakeItem(times):
    session = db.session
    list_revision = ['A','A0','B','C']
    list_stage = ['Y','Y1','Y2','YF']
    for n in range(times):
        for revision in list_revision:
            for stage in list_stage:
                
                os.chdir(UPLOAD_FOLDER + "/fakeDras")
                for file in glob.glob("*.xlsx"):
                    
                    fakedoc = fakeDoc(revision,stage)
                    print('Fake DOC:', fakedoc) 
                    csid = str(uuid.uuid4())
                    fakename =UPLOAD_FOLDER+ "/fakeDrasOUT/" + csid +"_sep_" + fakedoc
                    fakefile = "fakeDrasOUT/" + csid +"_sep_" + fakedoc 
                    
                    copyfile(file, fakename)
                    
                    cs = Drascommentsheet(cs_file=fakefile,
                            current=True,
                            created_by_fk='1',
                            changed_by_fk='1',

                            documentClientCode = 'Any Document Client Code',
                            issuetype_id= '1',
                            actionrequired_id='1',
                            notificationItem='Email/Transmittal Id',

                            actualDate=gen_datetime(),
                            expectedDate=gen_datetime(),
                            plannedDate=gen_datetime(),)
                    try:
                        data = get_fake_data_from_cs(cs) 
                        session.add(cs)  
                    except:
                        pass
    session.commit()

def fakeItem2(serials):
    session = db.session
    units = session.query(Unitmodel).all()
    doctypes = ['MOM','PID','MR','SOW']
    revisions = ['A','B','C']
    stages = ['Y','Y1','Y2','YF']
    sheet = '01'
    
    for unit in units:
        for doctype in doctypes:
            for revision in revisions:
                for stage in stages:
                    for serial in serials:
                        doc = "-".join(('DRAS_2544',
                                        unit.code,
                                        doctype,
                                        serial,
                                        sheet + "_" + revision + stage+ ".xlsx"))
            
                        os.chdir(UPLOAD_FOLDER + "/fakeDras")
                        for file in glob.glob("*.xlsx"):
                            
                            fakedoc = doc
                            print('Fake DOC:', fakedoc) 
                            csid = str(uuid.uuid4())
                            fakename =UPLOAD_FOLDER+ "/fakeDrasOUT/" + csid +"_sep_" + fakedoc
                            fakefile = "fakeDrasOUT/" + csid +"_sep_" + fakedoc 
                            
                            copyfile(file, fakename)
                           
                            cs = Drascommentsheet(cs_file=fakefile,
                                    current=True,
                                    created_by_fk='1',
                                    changed_by_fk='1',

                                    documentClientCode = 'Any Document Client Code',
                                    issuetype_id= '1',
                                    actionrequired_id='1',
                                    notificationItem='Email/Transmittal Id',

                                    actualDate=gen_datetime(),
                                    expectedDate=gen_datetime(),
                                    plannedDate=gen_datetime(),)
                            try:
                                data = get_fake_data_from_cs(cs) 
                                session.add(cs)  
                            except:
                                pass
    

    session.commit()


from tempfile import NamedTemporaryFile

def fakeItem3(times):
    session = db.session
    units = session.query(Unitmodel).all()
    doctypes = ['MOM','PID','MR','SOW']
    revisions = ['A','B']
    stages = ['Y']
    sheet = '01'
    

    def random_element(lista):
        return lista[randint(0,len(lista)-1)]
    
    for n in range(times):
        print('                --------------------------------|',n)
        c_unit = random_element(units)
        c_doctype = random_element(doctypes)
        serials = randint(0,9999)
        
        c_serial= str(serials).zfill(4)
        
        for revision in revisions:
            for stage in stages:
                doc = "-".join(('DRAS_2544',
                                        c_unit.code,
                                        c_doctype,
                                        c_serial,
                                        sheet + "_" + revision + stage+ ".xlsx"))
            
                #os.chdir(UPLOAD_FOLDER + "/fakeDras")
                
                f = NamedTemporaryFile()
                f.file = UPLOAD_FOLDER + "/fakeDras/DRAS_2544-13-MOM-4561-09_A0Y.xlsx"
                
                csid = str(uuid.uuid4())
                f.name = csid +"_sep_" + doc
                
                fakefile = f
                it_list = session.query(Drasissuetype).all()
                ac_list = session.query(Drasactionrequired).all()

                cs = Drascommentsheet(cs_file=fakefile.name,
                        current=True,
                        created_by_fk='1',
                        changed_by_fk='1',

                        documentClientCode = 'Any Document Client Code',
                        issuetype_id= random.choice(it_list).id,
                        actionrequired_id=random.choice(ac_list).id,
                        notificationItem='Email/Transmittal Id',

                        actualDate=gen_datetime(),
                        expectedDate=gen_datetime(),
                        plannedDate=gen_datetime(),)
                    
                data = get_fake_data_from_cs2(cs) 
                session.add(cs)
                f.close()    
                       
    session.commit()
    
'''
#fakeItem(1)                  
#fakeItem2(['0020','0054','0124','0155','0034'])
#fakeItem2(['0020','0054'])
#fakeItem3(99)
 
 

#fakeItem3(40)