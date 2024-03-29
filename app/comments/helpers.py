
from config import UPLOAD_FOLDER
import openpyxl
from app.models import Drasdocument, Drasrevision, Drascommentsheet, Drascomment, Splitofworks, Unitmodel, Disciplinedras, Tagdiscipline
from datetime import datetime
from flask import flash, abort
from app import db
from random import randint
from datetime import timedelta


def date_parse(field):
    if isinstance(field, datetime):
        #print('We got a valid Date!')
        return field
    try:
        return datetime.strptime(field,'%d/%m/%y')
    except:
        #print('Date Parse ERROR --------------- GOT:', field)
        return None

def check_labels(item):
    # 
    # Check File Labels
    # 
    #  
    item = item.cs_file
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item)
    csSheet = csFile.active
    
    # Duplicate key (label in excel) for check purpose
    header_labels_dict = {
        'Reference' : csSheet['C8'].value,
        'Date' : csSheet['D8'].value,

        'Reference' : csSheet['G8'].value,
        'Date' : csSheet['G9'].value,
        'Material requisition': csSheet['G10'].value,
        'Vendor Name' : csSheet['G11'].value,

        'Rev.' : csSheet['K9'].value,
        'Description': csSheet['K10'].value,
        'Issued by (Contractor Discipline)': csSheet['K11'].value
    }

    for key, value in header_labels_dict.items():
        #print(key, value)
        if key != value:
            #flash(('Header Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            abort(400,('Header Label ' + key + ' Not Found, Check your DRAS Template!'))
    
    table_label_dict = {
        'Pos.' : csSheet['B14'].value,
        
        'Status' : csSheet['G16'].value,

        'Date' : csSheet['L15'].value,

    }

    for key, value in table_label_dict.items():
        #print(key, value)
        if key != value:
            flash(('Table Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            print(('Table Label ' + key + ' Not Found, Check your DRAS Template!'))
            return False
    
    print('-------------------- CHECK FUNCTION TRUE')
    return True

def update_data_from_cs(item):
    session = db.session   
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.drasdocument
        
    except:
        return abort(400, 'Error in your File Name.')
    #doc = session.query(Document).filter(Document.id == item.document_id).first()

    #rev = session.query(Drasrevision).filter(Drasrevision.id == item.drasrevision_id).first()
    rev = item.drasrevision
        
    rev.stage = item.stage
    
    print(rev.id)
    #print(doc.id)
    #session.flush()
    #print(rev.id)
    #print(doc.id)

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    
    print(csSheet['C9'].value)
    
    try:
        #print('before maybe here')
        #item.revision_id = rev.id
        #item.document_id = item.document_id
        
        
        item.ownerTransmittalReference = csSheet['C9'].value
        
        item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
        print(csSheet['C12'].value)
        item.response_status = csSheet['C12'].value
        
        
        item.contractorTransmittalReference = csSheet['H8'].value
        item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
        

        item.contractorTransmittalMr = csSheet['H10'].value
        item.contractorTransmittalVendor = csSheet['H11'].value
        
        item.documentReferenceDoc = csSheet['L8'].value
        item.documentReferenceRev = csSheet['L9'].value
        item.documentReferenceDesc = csSheet['L10'].value
        item.documentReferenceBy = csSheet['L11'].value
        
        '''
        BODY - CREATE NEW COMMENTS FOR THIS CS
        '''
        
        

        #doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
        session.query(Drascomment).filter(Drascomment.drasdocument_id == document.id).delete()
        #print('Document',doc.id, doc.name)
        cs_list = session.query(Drascommentsheet).filter( 
                                    Drascommentsheet.drasdocument_id == document.id,
                                    Drascommentsheet.current == True).all()
        #item.stage = rev_stage
        for cs in cs_list:  
            cs.current = False
            print('cs', cs.current, cs.id, 'item',item.current, item.id)
        item.current = True
        
        print('Item Current', item.current)
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            
            
            

            if row[0].value is not None:
                #print(row[0].value) 
                comment = Drascomment(

                    drasrevision_id = rev.id,
                    drascommentsheet = item,

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                    ownerCounterReplyComment = row[7].value,

                    finalAgreementDate = date_parse(csSheet['L15'].value),
                    finalAgreemntCommentDate = date_parse(row[8].value),
                    finalAgreementComment = row[9].value,

                    commentStatus = row[10].value,
                )
                print('-----         ************       --------')
                print(item.current)
                if item.current == True:
                    
                    comment.drasdocument_id = document.id
                    
                    print(comment.drasdocument_id, document.id)
                    

                
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        
        print('maybe here')
        session.commit()
        return True
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')


def get_oc(unit, discipline):
    session = db.session
    unit_id = session.query(Unitmodel).filter(Unitmodel.code == unit).first()
    discipline_id = session.query(Disciplinedras).filter(Disciplinedras.name == discipline).first()
    '''
    if unit_id and discipline_id:
        splitOfWorks = session.query(Splitofworks).filter(
                    Splitofworks.unit_id == unit_id.id,
                    Splitofworks.discipline_id == discipline_id.id).first()
        print(' ------- ----- ----  GET OC by UNIT and DISCIPLINE ----------')

        print(unit_id.moc_id, splitOfWorks.oc_id)
        return unit_id.moc_id, splitOfWorks.oc_id
    '''
    if unit_id:
        print(' ------- ----- ----  GET OC ONLY by UNIT --- NO DISCIPLINE ----------')
        return unit_id.moc_id, unit_id.dedoc_id
    return abort(400,'Unit not Found, check your file name.')

 
def get_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[-1].split('.')[0]
        print('Heeeeeeeeeeeere ********************' )
        
        try:
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
        except:
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Error in file name. Check Your File!')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        #fake Discipline
        
        discipline = csSheet['L11'].value

        print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
        session.add(doc)
        print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id
    
    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision, document)
    print('found', rev)
    if rev is None:
        print(rev)
        print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc)
        session.add(rev)
        
    rev.stage = rev_stage
    
     
    session.flush()

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['L8'].value
    item.documentReferenceRev = csSheet['L9'].value
    item.documentReferenceDesc = csSheet['L10'].value
    
    # Discipline
    item.documentReferenceBy = csSheet['L11'].value

    #item.documentReferenceBy = fdiscipline
    indoor = ['Y','Y2']
    outdoor = ['S','Y1','Y3']

    if rev_stage in indoor: 
        item.expectedDate = item.actualDate + timedelta(days=7)
    if rev_stage in outdoor:
        item.expectedDate = item.actualDate + timedelta(days=14)
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
    

    try:
        none_count = 0
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value), none_count)
            
            if row[0].value is not None and row[1].value is not None:
                print(row[0].value)
                #  
                comment = Drascomment(
                    drasrevision_id = rev.id,
                    drascommentsheet = item,
                    tagdiscipline= session.query(Tagdiscipline).filter(
                                                Tagdiscipline.start <= int(row[1].value), 
                                                Tagdiscipline.finish >= int(row[1].value)).first(), 

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                    ownerCounterReplyComment = row[8].value,

                    finalAgreementDate = date_parse(csSheet['M15'].value),
                    finalAgreemntCommentDate = date_parse(row[10].value),
                    finalAgreementComment = row[11].value,

                    commentStatus = str(row[12].value),
                )

                if item.current:
                    print('BLOCKED HERE ------------------ //////////////////////')
  
                    
                    comment.drasdocument_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                session.add(comment)
            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 5:
                    raise Exception('Excel BAD FORMAT: Infinite Comments')
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()

        session.commit()
        return doc.id

    
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')
     

def get_fake_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[1].split('.')[0]
        revision = full_revision[:full_revision.index('Y')]
        rev_stage = full_revision[full_revision.index('Y'):]

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Error in file name. Check Your File!')

    doc = session.query(Document).filter(Document.name == document).first()
    
    if doc is None:
        
        discipline = csSheet['K11'].value
        print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Document(name=document, 
                        moc_id=moc, 
                        dedoc_id=dedoc,
                        created_by_fk='1',
                        changed_by_fk='1')
        session.add(doc)
        session.flush()
        print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id

    rev = session.query(Revision).filter(Revision.name == revision, Revision.document_id == doc.id).first() 
    print('searching for revision, document:', revision,rev_stage, document)
    print('found', rev)
    if rev is None:
        print(rev)
        print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Revision(name=revision, document=doc,created_by_fk='1',
                        changed_by_fk='1')
        rev.stage = rev_stage
        session.add(rev)
        session.flush()
        
    
    
    
    
    

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.revision_id = rev.id
    item.document_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['L8'].value
    item.documentReferenceRev = csSheet['L9'].value
    item.documentReferenceDesc = csSheet['L10'].value
    item.documentReferenceBy = csSheet['L11'].value
    
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Comment).filter(Comment.document_id == doc.id).delete()
        
        commentSheets = session.query(Commentsheet).filter(Commentsheet.document_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
            cs.changed_by_fk = '1'

    try:
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            #print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            if row[0].value is not None:
                #print(row[0].value) 
                comment = Comment(
                    revision_id = rev.id,
                    commentsheet = item,

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                    ownerCounterReplyComment = row[7].value,

                    finalAgreementDate = date_parse(csSheet['L15'].value),
                    finalAgreemntCommentDate = date_parse(row[9].value),
                    finalAgreementComment = row[10].value,

                    commentStatus = str(row[11].value),
                    
                    created_by_fk='1',
                    changed_by_fk='1'
                )
                if item.current:  
                    
                    comment.document_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()

        session.commit()
        return doc.id

    
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')
     

def get_fake_data_from_cs2(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+'fakeDras/DRAS_2544-13-MOM-4561-09_A0Y.xlsx', data_only=True)
    csSheet = csFile.active 
 
    #print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try: 
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[1].split('.')[0]
        revision = full_revision[:full_revision.index('Y')]
        rev_stage = full_revision[full_revision.index('Y'):]

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Error in file name. Check Your File!')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        
        discipline = csSheet['K11'].value
        #print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, 
                        moc_id=moc, 
                        dedoc_id=dedoc,
                        created_by_fk='1',
                        changed_by_fk='1')
        session.add(doc)
        session.flush()
        #print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id

    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision,rev_stage, document)
    #print('found', rev)
    #rev.stage = rev_stage
    #rev.changed_by_fk = '1'
    if rev:
        rev.stage = rev_stage
        rev.changed_by_fk = '1'
    else:
        #print(rev)
        #print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc,
                created_by_fk='1',
                changed_by_fk='1',
                stage=rev_stage)
        
        session.add(rev)
        session.flush()
    
    #rev.stage = rev_stage
    #rev.changed_by_fk = '1'
    
    
    
    
    

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['L8'].value
    item.documentReferenceRev = csSheet['L9'].value
    item.documentReferenceDesc = csSheet['L10'].value
    item.documentReferenceBy = csSheet['L11'].value
    
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
            cs.changed_by_fk = '1'
    
    # random comment status
    #
    #   
    
    
    def random_status():
        status = ['Open', 'Closed']
        return status[randint(0,1)]
        
    try:
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            #print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            if row[0].value is not None:
                #print(row[0].value) 
                comment = Drascomment(
                    drasrevision_id = rev.id,
                    drascommentsheet = item,

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                    ownerCounterReplyComment = row[7].value,

                    finalAgreementDate = date_parse(csSheet['L15'].value),
                    finalAgreemntCommentDate = date_parse(row[9].value),
                    finalAgreementComment = row[10].value,

                    #commentStatus = str(row[11].value),
                    commentStatus = random_status(),

                    created_by_fk='1',
                    changed_by_fk='1'
                )
                if item.current:  
                    
                    comment.drasdocument_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()

        session.commit()
        return doc.id

    
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')
     

def text_decode(comment):
    ''' Decode ASCII in comments'''
    if comment:
        string_encode = str(comment).encode("ascii", "ignore")
        string_decode = string_encode.decode()
        
        return string_decode
    return comment


'''
def find_rev():
    session = db.session
    document = '2544-17-DW-0510-04'
    revision = 'CYF'

    rev = session.query(Revision).filter(Revision.name == revision, Document.name == document).first()
    if rev is None:
        print('     ----- Rev is NONE')
    print('    ------  RR rev: ',rev.id, rev.name, rev.document)
    
    return
'''
#find_rev()

def set_current_last_actual_date():
    session = db.session
    docs = session.query(Drasdocument).all()
    bad_file = []
    
    for doc in docs:
        try:
            cs_list = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id
            ).all()
            print(cs_list)
            for c in cs_list:
                print(c.id)
                c.changed_by_fk = '1'
                c.created_by_fk = '1'
                c.current = False
            session.commit()    
                
            
            cs = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id
            ).order_by(Drascommentsheet.actualDate.desc()).first()
            print(doc,cs.id)
            cs.current = True
            cs.changed_by_fk = '1'
            cs.created_by_fk = '1'
        except:
            bad_file.append(cs)      
            
    session.commit()

#set_current_last_actual_date() 
    # 
    #  
def set_expected_date():
    session = db.session
    cs_list = session.query(Drascommentsheet).all()
    indoor = ['Y','Y2']
    outdoor = ['Y1', 'Y3','S']
    for cs in cs_list:
        try:

            print(cs.id)
            if cs.actualDate:
                if cs.stage in indoor:
                    cs.expectedDate = cs.actualDate + timedelta(days=7)
                elif cs.stage in outdoor:
                    cs.expectedDate = cs.actualDate + timedelta(days=15)
            cs.changed_by_fk = '1'
            cs.created_by_fk = '1'
        except:
            print('SOMETHING WRONG')
        
           
    session.commit()

#set_expected_date() 
        



import os
from random import choice, randrange
from time import sleep
def check_labels2(item):
    # 
    # Check File Labels
    # 
    #  
    #item = item.cs_file
    csFile = openpyxl.load_workbook(item)
    csSheet = csFile.active
    
    # Duplicate key (label in excel) for check purpose
    header_labels_dict = {
        'Reference' : csSheet['C8'].value,
        'Date' : csSheet['D8'].value,

        'Reference' : csSheet['G8'].value,
        'Date' : csSheet['G9'].value,
        'Material requisition': csSheet['G10'].value,
        'Vendor Name' : csSheet['G11'].value,

        'Rev.' : csSheet['K9'].value,
        'Description': csSheet['K10'].value,
        'Issued by (Contractor Discipline)': csSheet['K11'].value
    }

    for key, value in header_labels_dict.items():
        #print(key, value)
        if key != value:
            #flash(('Header Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            abort(400,('Header Label ' + key + ' Not Found, Check your DRAS Template!'))
    
    table_label_dict = {
        'Pos.' : csSheet['B14'].value,
        
        'Status' : csSheet['G16'].value,

        'Date' : csSheet['L15'].value,

    }

    for key, value in table_label_dict.items():
        #print(key, value)
        if key != value:
            flash(('Table Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            print(('Table Label ' + key + ' Not Found, Check your DRAS Template!'))
            return False
    
    print('-------------------- CHECK FUNCTION TRUE')
    return True

def cs_data_report(dras_file):
        
    print(dras_file)
    cs_file = dras_file.split('/')[-1]
    print(cs_file)
    try:
        print('BEFORE ********************' )
        document = cs_file.split('DRAS_')[1].split('_')[0]
        

        full_revision = cs_file.split('DRAS_')[1].split('_')[1].split('.')[0]
        
        
        try:
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
        except:
            
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]
        
        oc_unit = document.split('-')[1]
        project = document.split('-')[0]
        print(document,revision,rev_stage,oc_unit,project)

    except:
        abort(400, 'Error in file name. Check Your File!')


    
    issue_type = 'it'
    action = 'ac'
    not_item = 'ni'
    actual_date = 'ad'
    
    return issue_type, action, not_item, actual_date

def batch_upload():
    session = db.session
    users = [1,2,3]
    batch_folder = 'init/dras_s'
    #batch_folder = 'settimo_batch'
    path = UPLOAD_FOLDER + batch_folder

    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))

    bad_file = []
    for f in files:
        try:
            print(f)
            
            '''
            issue_type, action, not_item, actual_date = cs_data_report(f)
        
            print(issue_type, action, not_item, actual_date)
            print(' BEFORE LABELS ********************' )
            check_labels2(f)
            print(' LABELS ********************' )
            '''
            csFile = openpyxl.load_workbook(f)
            csSheet = csFile.active

            notification_item = csSheet['H8'].value
            actual_date = csSheet['H9'].value

            full_file = f.split('/')[-1]
            document = full_file.split('_')[1]
            revision = full_file.split('_')[2].split('.')[0]

            print(document, revision)
            oc_unit = document.split('-')[1]
            project = document.split('-')[0]

            rev_stage = 'S'
            doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()

            if doc is None:
                #fake Discipline
                
                discipline = csSheet['L11'].value

                print('DOC is NONE *-----------           ************')
                moc, dedoc = get_oc(oc_unit, discipline)
                print('MOC - DEDOC ', moc, dedoc)
                doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
                doc.created_by_fk = '1'
                doc.changed_by_fk = '1'
                print('DOC           ************')
                session.add(doc)
                print('DOC  ADD         ************')
                session.flush()
                print('DOC  FLUSH         ************')
                print('Document',doc.name)
                print(revision, doc.id)
            # session flush for doc id
            session.flush()
            print('AFTER FLUSH')
            print(revision, doc.id)
            # search the same rev for this document by doc id
            
            print('BLOCKED HERE ------------------ //////////////////////')
            rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
            print('searching for revision, document:', revision, document)
            print('found', rev)
            if rev is None:
                print(rev)
                print('    ----------     Rev is None: ', revision, rev_stage, document)
                rev = Drasrevision(name=revision, drasdocument=doc, changed_by_fk='1',created_by_fk = '1')
                
                session.add(rev)
                #session.flush()
            
            rev.changed_by_fk = '1'
            rev.stage = rev_stage
            #rev.created_by_fk = '1'
            
            
            
            session.flush()
            

            '''
                HEADER - UPDATE THE COMMENT SHEET
            '''

                

            '''
            except:
                print('FAILED XXXXXXXXXXXXXXXXXXXXXXXXXX ',f)
                bad_file.append(f)
            '''

            #issue_type, action, not_item, actual_date = cs_data_report(f)
            user = choice(users)
            if isinstance(actual_date,str):
                try:
                    actual_date = datetime.strptime(actual_date,'%d/%m/%Y')
                except:
                    print('FAILED######################### ',f)
                    bad_file.append(f)
            
            drascs = Drascommentsheet(
                created_by_fk = user,
                changed_by_fk = user,
                cs_file = f,
                #issuetype = issue_type,
                #actionrequired = action,
                actualDate = actual_date,
                #actualDate = actual_date,
                notificationItem = notification_item,
                stage = 'S'
            )
            drascs.drasrevision_id = rev.id
            drascs.drasdocument_id = doc.id
            drascs.created_by_fk = '1'
            drascs.changed_by_fk = '1'
            session.add(drascs)
        except:
            bad_file.append(f)
    session.commit()
    print('***********BAD FILES LIST',len(bad_file))
    print(bad_file)

#batch_upload()  

def update_discipline():
    session = db.session
    users = [1,2,3]
    batch_folder = 'init/dras_s'
    #batch_folder = 'settimo_batch'
    path = UPLOAD_FOLDER + batch_folder

    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))

    bad_file = []
    for f in files:
        
        try:
            print(f)
            
            '''
            issue_type, action, not_item, actual_date = cs_data_report(f)
        
            print(issue_type, action, not_item, actual_date)
            print(' BEFORE LABELS ********************' )
            check_labels2(f)
            print(' LABELS ********************' )
            '''
            csFile = openpyxl.load_workbook(f)
            csSheet = csFile.active

            discipline = csSheet['L11'].value
            description = csSheet['L10'].value
            cs_rev = csSheet['L9'].value
            cs_doc = csSheet['L8'].value
            

            full_file = f.split('/')[-1]
            document = full_file.split('_')[1]
            revision = full_file.split('_')[2].split('.')[0]
            stage = 'S'

            print(document, revision)

            doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
            print('DOC ID', doc.id)
            rev = session.query(Drasrevision).filter(Drasrevision.drasdocument_id == doc.id,
                                Drasrevision.name == revision).first()
            
            print('FOUND +++++ =',doc.name, rev.name)
            cs = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id,
                Drascommentsheet.drasrevision_id == rev.id,
                Drascommentsheet.stage == stage
            ).first()
            cs.documentReferenceBy = discipline[:49]
            cs.documentReferenceDesc = description
            cs.documentReferenceRev = cs_rev
            cs.documentReferenceDoc = cs_doc

            cs.changed_by_fk = '1'
            print(cs.id,cs.documentReferenceBy,cs.documentReferenceDoc,cs.documentReferenceRev)
            session.commit()
        except:
            print('Something Wrong')
        
        
    
    

#update_discipline() 

from app.models import Drasvendor, Drasmr
 
def get_vendor_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    field_set = ['L8','L11','H10','H11']
    try:
        for field in field_set:
            len(csSheet[field].value)
    except:
        abort(400,'Controllare le celle L8 L11 H10 H11 nel DRAS.')
    
    # Controlla se questa MR ha lo stesso fornitore.
    #mr = session.query(Drasmr).filter(Drasmr.name == csSheet['H10'].value).first()
    #if mr and str(mr.drasvendor) != csSheet['H11'].value: 
        #abort(400,'MR assegnata ad un altro fornitore o il nome del Vendor (cella H11) non corrisponde.')

    
    try:
        #document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        document = csSheet['L8'].value
        document_fullname = item.cs_file.split('_sep_DRAS_')[1]
        try:
            document_filename = document_fullname[:document_fullname.rindex('_')]  
        except Exception as e:
            print(str(e))
        #document_filename = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[-1].split('.')[0] 
        print('Heeeeeeeeeeeere -------------',document_filename,document )
        
        if document != document_filename:
            print('++++ + + + + +++ WRONG DRAS DOCUMENT')
            flash('Il documento estrapolato dal Filename e il documento riportato nel DRAS (cella H11) sono diversi. FILENAME: '+ document_filename + ' -> DRAS: ' + document, 'info')
            abort(400,'Documento nel DRAS errato')

        #print('full revision', full_revision)
        
        try:
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
            print('S rev_stage', rev_stage)
        except:
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]
            print('Y rev_stage', rev_stage) 
        try:
            oc_unit = csSheet['H10'].value.split('-')[1]
        except:
            print('H10 cell (MR) is none or blocked')
            flash('H10 cell (MR) is none or blocked', category='info')

            abort(409)
        #print(' -----  OC unit  ---- - - - - ',csSheet['L8'].value.split('-')[1] )
        project = '2544' 

    

        doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
        if doc:
            print('Doc is there', doc)
        if doc is None:
            #fake Discipline
            
            discipline = csSheet['L11'].value

            print('DOC is NONE *-----------           ************')
            moc, dedoc = get_oc(oc_unit, discipline)
            print('MOC - DEDOC ', moc, dedoc)
            doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
            session.add(doc)
            print('Document',doc.name)

        # session flush for doc id
        # search the same rev for this document by doc id
        
        rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
        print('searching for revision, document:', revision, document)
        print('found', rev) 
        if rev is None:
            print(rev)
            print('    ----------     Rev is None: ', revision, rev_stage, document)
            rev = Drasrevision(name=revision, drasdocument=doc)
            session.add(rev)
            
        rev.stage = rev_stage
        #Check If this DRAS STAGE already exist
        ds = session.query(Drascommentsheet).filter(
            Drascommentsheet.drasdocument_id == doc.id,
            Drascommentsheet.drasrevision_id == rev.id,
            Drascommentsheet.stage == rev_stage
        ).first()
        if ds:
            print(doc.id, rev.id, rev_stage)
            flash('This DRAS Stage already Exist', category='info')

            abort(409)


        print('before flush')
        session.flush()
        print('after flush') 

        #
        #  SET VENDOR AND MR RELATION
        #
        
        vendor = session.query(Drasvendor).filter(Drasvendor.name == csSheet['H11'].value).first()
        mr = session.query(Drasmr).filter(Drasmr.name == csSheet['H10'].value).first()
        if vendor is None:
                
            vendor = Drasvendor(
                name = csSheet['H11'].value,
                created_by_fk = '1',
                changed_by_fk = '1',
            )
        if mr is None:
            
            mr = Drasmr(
                name = csSheet['H10'].value,
                created_by_fk = '1',
                changed_by_fk = '1',
                drasvendor = vendor
            )

        #
        #    HEADER - UPDATE THE COMMENT SHEET
        #
        print('++++ debug revi id +++++', rev.id)
        item.drasrevision_id = rev.id
        item.drasdocument_id = doc.id
        item.drasvendor = vendor
        item.drasmr = mr
        
        if csSheet['L14'].value == 'Final Agreement':

            item.ownerTransmittalReference = csSheet['C9'].value
            item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
            item.response_status = csSheet['C12'].value

            item.contractorTransmittalReference = csSheet['H8'].value
            item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
            item.contractorTransmittalMr = csSheet['H10'].value
            item.contractorTransmittalVendor = csSheet['H11'].value

            item.documentReferenceDoc = csSheet['L8'].value
            item.documentReferenceRev = csSheet['L9'].value
            item.documentReferenceDesc = csSheet['L10'].value
            
            # Discipline
            item.documentReferenceBy = csSheet['L11'].value
            
        else:
            item.ownerTransmittalReference = csSheet['C9'].value
            item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
            item.response_status = csSheet['C12'].value

            item.contractorTransmittalReference = csSheet['H8'].value
            item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
            item.contractorTransmittalMr = csSheet['H10'].value
            item.contractorTransmittalVendor = csSheet['H11'].value

            item.documentReferenceDoc = csSheet['K8'].value
            item.documentReferenceRev = csSheet['K9'].value
            item.documentReferenceDesc = csSheet['K10'].value
            
            # Discipline
            item.documentReferenceBy = csSheet['K11'].value

        #
        # SET EXPECTED DATE BASED ON STAGE
        #

        indoor = ['S','Y','Y2']
        outdoor = ['Y1','Y3']

        if rev_stage in indoor: 
            item.expectedDate = item.actualDate + timedelta(days=14)
        if rev_stage in outdoor:
            item.expectedDate = item.actualDate + timedelta(days=7)
        
        #
        # SET DRAS AS CURRENT
        #
        '''
        if item.current:
            session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
            
            commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
            item.stage = rev_stage

            for cs in commentSheets:
                cs.current = False
        else:
            item.stage = rev_stage
        '''
        item.stage = rev_stage
        session.commit()
    except:
        abort(400, 'DRAS Error: controllare il nome del file (revision e stage dopo _ ) e nel DRAS i campi Vendor, Material R e Document. Se presenti, verificare lo Split of Works per questa Unit. ')    
    
    
    #
    # BODY - CREATE NEW COMMENTS FOR THIS CS
    #   

    try:
        none_count = 0
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value), none_count)
            
            if row[0].value is not None and row[1].value is not None:
                print(row[0].value)
                
                if csSheet['L14'].value == 'Final Agreement':
                    comment = Drascomment(
                        drasdocument_id = doc.id,
                        drasrevision_id = rev.id,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 
                    

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value) ,

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                        ownerCounterReplyComment = text_decode(row[8].value),

                        finalAgreementDate = date_parse(csSheet['M15'].value),
                        finalAgreemntCommentDate = date_parse(row[10].value),
                        finalAgreementComment = text_decode(row[11].value),

                        commentStatus = str(row[12].value),
                    )

                    #if item.current:
                        #comment.drasdocument_id = doc.id
                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                    #session.add(comment)
                    #session.commit()
                else:
                    comment = Drascomment(
                        drasdocument_id = doc.id,
                        drasrevision_id = rev.id,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                        ownerCounterReplyComment = text_decode(row[7].value),

                        finalAgreementDate = date_parse(csSheet['L15'].value),
                        finalAgreemntCommentDate = date_parse(row[9].value),
                        finalAgreementComment = text_decode(row[10].value),

                        commentStatus = str(row[11].value),
                        
                    )
                ## SET MAX LENGHT
                if comment.commentStatus:
                    comment.commentStatus = comment.commentStatus[:20]
                if comment.ownerCommentBy:
                    comment.ownerCommentBy = comment.ownerCommentBy[:50]
                
                if comment.pos:
                    comment.pos = str(comment.pos)[:5]

                if comment.contractorReplyStatus:
                    comment.contractorReplyStatus = comment.contractorReplyStatus[:100]
                
                session.add(comment)

            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 35:
                    break
                    #raise Exception('Excel BAD FORMAT: Infinite Comments')
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        item.note = 'Piero003'
        session.commit()
        return doc.id 
    except:
        session.rollback()
        
        return flash('COMMENTS ERROR 003 | Impossibile caricare i commenti per questo DRAS', category='warning')
        #item.note = 'COMMENTS ERROR 003: Badly Formatted. Please find the attached original DRAS in order to review you comments.'

        


def cs_sanification():
    print('** *-* * --* -* -* - *- * -* -* -*-*- *- *-*** -**--- **- *- *')
    session = db.session
    cs_list = session.query(Drascommentsheet).filter(Drascommentsheet.documentReferenceBy=="Issued by (Contractor Discipline)").all()
    print('CS da sanare:',len(cs_list))
    wrong_cs = []
    
    for cs in cs_list:
        try:
            csFile = openpyxl.load_workbook(UPLOAD_FOLDER+cs.cs_file, data_only=True)
            csSheet = csFile.active
            cs.documentReferenceBy = csSheet['L11'].value
            cs.documentReferenceDesc = csSheet['L10'].value
            cs.documentReferenceRev = csSheet['L9'].value
            cs.documentReferenceDoc = csSheet['L8'].value
            cs.changed_by_fk = '1'
        except:
            print('CS FAIL - ID:', cs.id)
            wrong_cs.append(cs.id)
            pass
    
    session.commit()
    print('CS NOT FOUND LIST')
    print(wrong_cs)


def update_vendor_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
    
    session.query(Drascomment).filter(Drascomment.drascommentsheet_id == item.id).delete()
    
    print('Update Comments -+-+-+')

    try:
        none_count = 0
        print('DRAS', item.drasdocument, item.drasrevision, item.stage)
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            
            
            if row[0].value is not None and row[1].value is not None:
                print('pos',row[0].value,'tag',row[1].value,'info',row[2].value)
                
                if csSheet['L14'].value == 'Final Agreement':
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 
                    

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value) ,

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                        ownerCounterReplyComment = text_decode(row[8].value),

                        finalAgreementDate = date_parse(csSheet['M15'].value),
                        finalAgreemntCommentDate = date_parse(row[10].value),
                        finalAgreementComment = text_decode(row[11].value),

                        commentStatus = str(row[12].value),
                    )

                    #if item.current:
                        #comment.drasdocument_id = doc.id
                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                    #session.add(comment)
                    #session.commit()
                else:
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                        ownerCounterReplyComment = text_decode(row[7].value),

                        finalAgreementDate = date_parse(csSheet['L15'].value),
                        finalAgreemntCommentDate = date_parse(row[9].value),
                        finalAgreementComment = text_decode(row[10].value),

                        commentStatus = str(row[11].value),
                        
                    )
                ## SET MAX LENGHT
                if comment.commentStatus:
                    comment.commentStatus = comment.commentStatus[:20]
                if comment.ownerCommentBy:
                    comment.ownerCommentBy = comment.ownerCommentBy[:50]
                
                if comment.pos:
                    comment.pos = str(comment.pos)[:5]

                if comment.contractorReplyStatus:
                    comment.contractorReplyStatus = comment.contractorReplyStatus[:100]
                
                session.add(comment)

            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 35:
                    break
                    #raise Exception('Excel BAD FORMAT: Infinite Comments')
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        item.note = 'Piero003'
        session.commit()
        return item.drasdocument_id 
    except Exception as e:
        session.rollback()
        print(str(e))
        return flash('COMMENTS ERROR 003 | Impossibile caricare i commenti per questo DRAS', category='warning')
        #item.note = 'COMMENTS ERROR 003: Badly Formatted. Please find the attached original DRAS in order to review you comments.'




def update_all_vendor_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
    
    deleted = session.query(Drascomment).filter(Drascomment.drascommentsheet_id == item.id).delete()
     
    #print('Delete Comments -+-+-+',item,deleted) 

    try:
        none_count = 0
        comments = 0
        
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            
            
            if row[0].value is not None and row[1].value is not None:
                #print('pos',row[0].value,'tag',row[1].value,'info',row[2].value)
                
                if csSheet['L14'].value == 'Final Agreement':
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 
                    

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value) ,

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                        ownerCounterReplyComment = text_decode(row[8].value),

                        finalAgreementDate = date_parse(csSheet['M15'].value),
                        finalAgreemntCommentDate = date_parse(row[10].value),
                        finalAgreementComment = text_decode(row[11].value),

                        commentStatus = str(row[12].value),
                    )

                    #if item.current:
                        #comment.drasdocument_id = doc.id
                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                    #session.add(comment)
                    #session.commit()
                else:
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                        ownerCounterReplyComment = text_decode(row[7].value),

                        finalAgreementDate = date_parse(csSheet['L15'].value),
                        finalAgreemntCommentDate = date_parse(row[9].value),
                        finalAgreementComment = text_decode(row[10].value),

                        commentStatus = str(row[11].value),
                        
                    )
                ## SET MAX LENGHT
                if comment.commentStatus:
                    comment.commentStatus = comment.commentStatus[:20]
                if comment.ownerCommentBy:
                    comment.ownerCommentBy = comment.ownerCommentBy[:50]
                
                if comment.pos:
                    comment.pos = str(comment.pos)[:5]

                if comment.contractorReplyStatus:
                    comment.contractorReplyStatus = comment.contractorReplyStatus[:100]

                comment.created_by_fk = '1'
                comment.changed_by_fk = '1'
                
                comments += 1
                session.add(comment)

            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 35:
                    break
                    #raise Exception('Excel BAD FORMAT: Infinite Comments')
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        session.commit()
        
        return item, comments 
    except Exception as e:
        session.rollback()
        
        #return flash('COMMENTS ERROR 003 | Impossibile caricare i commenti per questo DRAS', category='warning')
        #item.note = 'COMMENTS ERROR 003: Badly Formatted. Please find the attached original DRAS in order to review you comments.'

#rtese uwsgi[2456]: DRAS DWG-19111.01-01 4X YF - DRAS QPE069-9002 BX YF
def update_all_YF():
    session = db.session
    cs_list = session.query(Drascommentsheet).filter(
                            Drascommentsheet.stage == 'YF',
                            Drascommentsheet.note == '').all() 

    cs_count = 0
    cs_len = len(cs_list)
    errors_list = []
    print('leng of cs_list',cs_len)
    for cs in cs_list[:1000]:
        #print(cs)
        cs_count += 1 
        comments = session.query(Drascomment).filter(Drascomment.drascommentsheet_id == cs.id).all()
        
        if comments:
            cs.note = 'Comments Found'   
            #print(cs_count,'/', cs_len,'CS: ',cs,'Comments Found:',len(comments))
            cs.changed_by_fk = '1'
            session.commit() 
            #print(cs, 'has comments', len(comments))
        else:
            try:
                doc,comments = update_all_vendor_data_from_cs(cs)
                cs.note = 'Comments Updated'   
                print(cs_count,'/', cs_len,'CS: ',cs,'Comments Updated:',comments)
                cs.changed_by_fk = '1'
                session.commit()  
            except Exception as e:
                print(cs, ' |+-+- error -+-+| ', str(e))
                cs.note = str(e) 
                cs.changed_by_fk = '1'
                errors_list.append((cs,str(e)))
                session.commit()
    print('ERRORS',errors_list)

 
def piero_items_u():
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    csFile = openpyxl.load_workbook('app/comments/Piero_vendor.xlsx', data_only=True)
    csSheet = csFile.active

    start = 0

    for row in csSheet.iter_rows(min_row=2, min_col=1):
        start += 1
        try:
            document = row[0].value[:str(row[0].value).rindex('_')]
            rev = row[0].value[str(row[0].value).rindex('_')+1:]
            
            dras_doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
            
            if dras_doc: 
                dras_rev = session.query(Drasrevision).filter(
                    Drasrevision.name == rev,
                    Drasrevision.drasdocument_id == dras_doc.id).first()

                if dras_rev:
                    dras = session.query(Drascommentsheet).filter(
                        Drascommentsheet.stage == 'YF',
                        Drascommentsheet.drasrevision_id == dras_rev.id).first() 
                    
                    if dras:
                        dras, comments = update_all_vendor_data_from_cs(dras)
                        dras.note = 'Piero003'
                        dras.changed_by_fk = '1'
                        session.commit()
                        print(start,dras.drasdocument, dras.drasrevision, dras.stage, 'C: ',comments)

                        csSheet.cell(row=row[0].row,column=3).value = 'OK'
                        csSheet.cell(row=row[0].row,column=4).value = comments 
                    else:
                        print(start,' ',document,'NOT FOUND')
                        csSheet.cell(row=row[0].row,column=3).value = 'NO DRAS'
                else:
                        csSheet.cell(row=row[0].row,column=3).value = 'NO REV'
                        print(start,' ',document,'NOT FOUND')
            else:
                        csSheet.cell(row=row[0].row,column=3).value = 'NO DOC' 
                        print(start,' ',document,'NOT FOUND') 

        except Exception as e:
            csSheet.cell(row=row[0].row,column=3).value = 'Error'
            csSheet.cell(row=row[0].row,column=5).value = str(e) 
            print(e) 
    
    csFile.save('Piero_vendor_check.xlsx')
    csFile.close() 


def piero_items():
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    csFile = openpyxl.load_workbook('app/comments/Piero_vendor.xlsx', data_only=True)
    csSheet = csFile.active

    start = 0 

    for row in csSheet.iter_rows(min_row=2, min_col=1):
        if row[2].value is None:
            start += 1
            try:
                document = row[0].value[:str(row[0].value).rindex('_')]
                rev = row[0].value[str(row[0].value).rindex('_')+1:]
                
                dras_doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
                
                if dras_doc: 
                    dras_rev = session.query(Drasrevision).filter(
                        Drasrevision.name == rev,
                        Drasrevision.drasdocument_id == dras_doc.id).first()

                    if dras_rev:
                        dras = session.query(Drascommentsheet).filter(
                            Drascommentsheet.stage == 'YF',
                            Drascommentsheet.drasrevision_id == dras_rev.id).first() 
                        
                        if dras:
                            dras, comments = update_all_vendor_data_from_cs(dras)
                            dras.note = 'Piero003'
                            dras.changed_by_fk = '1'
                            session.commit()
                            
                            print(start,dras.drasdocument, dras.drasrevision, dras.stage, 'C: ',comments)

                            csSheet.cell(row=row[0].row,column=3).value = 'OK'
                            csSheet.cell(row=row[0].row,column=4).value = comments 
                        else:
                            print(start,' ',document,'DRAS | NOT FOUND')
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DRAS'
                    else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO REV'
                            print(start,' ',document,'REV | NOT FOUND')
                else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DOC' 
                            print(start,' ',document,'DOC | NOT FOUND') 
                
                
            except Exception as e:
                csSheet.cell(row=row[0].row,column=3).value = 'Error'
                csSheet.cell(row=row[0].row,column=5).value = str(e) 
                print(e) 
    
    csFile.save('app/comments/Piero_vendor.xlsx')
    csFile.close() 


def check_comm_pos(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    try:
    
        csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
        csSheet = csFile.active
        
        #deleted = session.query(Drascomment).filter(Drascomment.drascommentsheet_id == item.id).delete()
        
        #print('Delete Comments -+-+-+',item,deleted) 

        error_pos = []
        
        for row in csSheet.iter_rows(min_row=17, max_row = 25, min_col=2):
            
            if row[0].value is None and row[1].value:
                return False
        return True
    except Exception as e:
        print(e)


def piero_pos_check_items():
    #CHECK COMMENTS POS
    session = db.session
    csFile = openpyxl.load_workbook('app/comments/Piero_vendor.xlsx', data_only=True)
    csSheet = csFile.active

    start = 0 

    for row in csSheet.iter_rows(min_row=2, min_col=1):
        if row[2].value == 'OK':
            start += 1
            try:
                document = row[0].value[:str(row[0].value).rindex('_')]
                rev = row[0].value[str(row[0].value).rindex('_')+1:]
                
                dras_doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
                
                if dras_doc: 
                    dras_rev = session.query(Drasrevision).filter(
                        Drasrevision.name == rev,
                        Drasrevision.drasdocument_id == dras_doc.id).first()

                    if dras_rev:
                        dras = session.query(Drascommentsheet).filter(
                            Drascommentsheet.stage == 'YF',
                            Drascommentsheet.drasrevision_id == dras_rev.id).first() 
                        
                        if dras:
                            result = check_comm_pos(dras)
                            
                            
                            print(start,dras.drasdocument, dras.drasrevision, dras.stage, 'C: ',result)

                            csSheet.cell(row=row[0].row,column=3).value = 'OK'
                            csSheet.cell(row=row[0].row,column=6).value = result 
                        else:
                            print(start,' ',document,'DRAS | NOT FOUND')
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DRAS'
                    else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO REV'
                            print(start,' ',document,'REV | NOT FOUND')
                else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DOC' 
                            print(start,' ',document,'DOC | NOT FOUND') 
                
                
            except Exception as e:
                csSheet.cell(row=row[0].row,column=3).value = 'Error'
                csSheet.cell(row=row[0].row,column=5).value = str(e) 
                print(e) 
    
    csFile.save('app/comments/Piero_vendor.xlsx')
    csFile.close() 